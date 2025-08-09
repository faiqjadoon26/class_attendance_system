import cv2
import os
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import simpledialog, messagebox, scrolledtext

# Constants
DATASET_PATH = 'dataset'
TRAINER_PATH = 'trainer/trainer.yml'
ATTENDANCE_PATH = 'attendance.xlsx'

# Load Haar Cascade for face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Initialize recognizer
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Load the trained model if it exists
if os.path.exists(TRAINER_PATH):
    recognizer.read(TRAINER_PATH)

name_labels = {}

class AttendanceApp:
    def __init__(self, root):
        self.root = root
        root.title("Face Recognition Attendance System")
        root.geometry("600x400")

        # Buttons
        btn_capture = tk.Button(root, text="Capture Faces", command=self.capture_faces)
        btn_capture.pack(pady=5)

        btn_train = tk.Button(root, text="Train Model", command=self.train_model)
        btn_train.pack(pady=5)

        btn_recognize = tk.Button(root, text="Recognize Faces", command=self.recognize_faces)
        btn_recognize.pack(pady=5)

        btn_remove = tk.Button(root, text="Remove Student", command=self.remove_student)
        btn_remove.pack(pady=5)

        btn_exit = tk.Button(root, text="Exit", command=root.quit)
        btn_exit.pack(pady=5)

        # ScrolledText for status messages
        self.log = scrolledtext.ScrolledText(root, width=70, height=10, state='disabled')
        self.log.pack(pady=10)

    def log_message(self, message):
        self.log.config(state='normal')
        self.log.insert(tk.END, message + '\n')
        self.log.see(tk.END)
        self.log.config(state='disabled')

    def capture_faces(self):
        name = simpledialog.askstring("Input", "Enter student name:", parent=self.root)
        if not name:
            self.log_message("[ERROR] No name entered.")
            return

        student_path = os.path.join(DATASET_PATH, name)
        os.makedirs(student_path, exist_ok=True)

        cap = cv2.VideoCapture(0)
        count = 0
        self.log_message(f"[INFO] Starting face capture for {name}...")

        while True:
            ret, frame = cap.read()
            if not ret:
                self.log_message("[ERROR] Failed to capture image.")
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.1, 4)

            for (x, y, w, h) in faces:
                count += 1
                face_img = gray[y:y + h, x:x + w]
                file_path = os.path.join(student_path, f"{name}_{count}.jpg")
                cv2.imwrite(file_path, face_img)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)

            cv2.imshow('Capture Faces - Press q to quit', frame)

            if cv2.waitKey(1) & 0xFF == ord('q') or count >= 20:
                break

        cap.release()
        cv2.destroyAllWindows()
        self.log_message(f"[INFO] Captured {count} face images for {name}.")

    def train_model(self):
        global name_labels
        faces = []
        labels = []
        label_id = 0

        if not os.path.exists(DATASET_PATH):
            self.log_message("[ERROR] Dataset folder not found.")
            return

        for root_dir, dirs, files in os.walk(DATASET_PATH):
            for name in dirs:
                name_labels[name] = label_id
                student_path = os.path.join(root_dir, name)
                for file in os.listdir(student_path):
                    if file.endswith("jpg"):
                        img_path = os.path.join(student_path, file)
                        img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
                        faces.append(np.array(img, 'uint8'))
                        labels.append(label_id)
                label_id += 1

        if len(faces) == 0:
            self.log_message("[ERROR] No faces found to train.")
            return

        recognizer.train(faces, np.array(labels))
        recognizer.save(TRAINER_PATH)
        self.log_message("[INFO] Training model completed.")
        self.log_message(f"[INFO] Name labels: {name_labels}")

    def mark_attendance(self, name, status):
        date_str = datetime.now().date().isoformat()
        time_str = datetime.now().time().strftime("%H:%M:%S")
        if not os.path.exists(ATTENDANCE_PATH):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Date", "Time", "Status"])
            workbook.save(ATTENDANCE_PATH)

        workbook = load_workbook(ATTENDANCE_PATH)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            if row[0] == name and row[1] == date_str:
                return True

        sheet.append([name, date_str, time_str, status])
        workbook.save(ATTENDANCE_PATH)
        return False

    def recognize_faces(self):
        if not os.path.exists(TRAINER_PATH):
            self.log_message("[ERROR] No trained model found. Please train the model first.")
            return

        cap = cv2.VideoCapture(0)
        self.log_message("[INFO] Starting face recognition. Press 'q' to quit.")

        while True:
            ret, frame = cap.read()
            if not ret:
                self.log_message("[ERROR] Failed to capture image.")
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.1, 4)

            for (x, y, w, h) in faces:
                face_img = gray[y:y + h, x:x + w]
                label, confidence = recognizer.predict(face_img)

                name = None
                for n, l in name_labels.items():
                    if l == label:
                        name = n
                        break

                if name:
                    self.log_message(f"[INFO] Recognized: {name} with confidence {confidence:.2f}")
                    if self.mark_attendance(name, 'Present'):
                        self.log_message(f"[INFO] {name} already marked present today.")
                        cap.release()
                        cv2.destroyAllWindows()
                        return
                else:
                    self.log_message("[INFO] Unknown face detected.")

            cv2.imshow('Recognize Faces - Press q to quit', frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()

    def remove_student(self):
        name = simpledialog.askstring("Input", "Enter the name of the student to remove:", parent=self.root)
        if not name:
            self.log_message("[ERROR] No name entered.")
            return

        student_path = os.path.join(DATASET_PATH, name)
        if os.path.exists(student_path):
            for file in os.listdir(student_path):
                file_path = os.path.join(student_path, file)
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            os.rmdir(student_path)
            self.log_message(f"[INFO] Removed student {name} and their data.")
        else:
            self.log_message("[ERROR] Student not found.")

def main():
    root = tk.Tk()
    app = AttendanceApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
